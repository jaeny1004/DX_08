from __future__ import annotations

from io import BytesIO
from typing import Any, Iterable, Mapping

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="D9EAD3")
TITLE_FILL = PatternFill("solid", fgColor="B6D7A8")
THIN_GRAY_FILL = PatternFill("solid", fgColor="F3F4F6")


REPORT_FIELD_LABELS: dict[str, str] = {
    "document_no": "문서번호",
    "file_name": "파일명",
    "year": "연도",
    "center_grid_id": "중심 격자 ID",
    "sido_name": "시도",
    "sigungu_name": "시군구",
    "block_grid_ids": "분석 블록 격자 ID",
    "center_annual_count": "중심 격자 연간 발생 이력 수",
    "center_cumulative_count": "중심 격자 누적 발생 이력 수",
    "block_annual_count": "블록 연간 발생 이력 수",
    "block_cumulative_count": "블록 누적 발생 이력 수",
    "risk_score": "위험도 점수",
    "risk_grade": "위험도 등급",
    "priority_score": "예찰 우선순위 점수",
    "priority_grade": "예찰 우선순위 등급",
    "field_report_id": "현장예찰 보고서 ID",
    "source_prediction_report_id": "연결 예측보고서 ID",
    "source_prediction_file": "연결 예측보고서 파일",
    "prediction_link_status": "예측보고서 연결 상태",
    "survey_datetime": "현장 예찰 일시",
    "surveyors": "예찰 담당자",
    "species": "확인 수종",
    "total_trees": "전체 확인 수목 수",
    "suspicious_count": "현장 이상징후 수",
    "sample_count": "시료 채취 수",
    "qr_code": "현장 QR 코드",
    "air_survey_plan_appendix": "항공예찰 계획 별지",
    "air_survey_result_appendix": "항공예찰 결과 별지",
    "appendix_status": "별지 입력 상태",
    "data_origin": "데이터 생성 기준",
    "source_field_file": "연결 현장예찰 보고서 파일",
    "link_status": "3종 보고서 연결 상태",
    "field_suspicious_count": "현장 이상징후 수",
    "control_status": "방제 검토 상태",
    "planned_count": "방제 검토 대상 수",
    "planned_shred_count": "파쇄 검토 수",
    "planned_fumigation_count": "훈증 검토 수",
    "planned_preventive_injection_count": "예방나무주사 검토 수",
    "planned_area_ha": "방제 검토 면적(ha)",
    "script_version": "생성기 버전",
}


REPORT_TYPE_LABELS = {
    "prediction": "발생 예측",
    "field_survey": "현장 예찰",
    "control": "방제",
}


def _display_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (list, tuple, set)):
        return " | ".join(str(item) for item in value)
    return value


def _style_header_row(ws, row_number: int) -> None:
    for cell in ws[row_number]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autosize_columns(ws, minimum: int = 12, maximum: int = 48) -> None:
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))

        ws.column_dimensions[column_letter].width = min(
            max(max_length + 2, minimum),
            maximum,
        )


def _add_detail_sheet(
    wb: Workbook,
    sheet_name: str,
    report_type: str,
    row: Mapping[str, Any],
) -> None:
    ws = wb.create_sheet(sheet_name)
    ws.freeze_panes = "A3"

    ws["A1"] = f"{REPORT_TYPE_LABELS.get(report_type, report_type)} 보고서 데이터"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].fill = TITLE_FILL
    ws.merge_cells("A1:B1")

    ws.append(["항목", "값"])
    _style_header_row(ws, 2)

    for key, value in row.items():
        ws.append([
            REPORT_FIELD_LABELS.get(key, key),
            _display_value(value),
        ])

    for cell in ws["A"]:
        cell.alignment = Alignment(vertical="top")
    for cell in ws["B"]:
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    _autosize_columns(ws)


def build_single_report_workbook(
    report_type: str,
    row: Mapping[str, Any],
) -> bytes:
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    _add_detail_sheet(
        wb=wb,
        sheet_name=REPORT_TYPE_LABELS.get(report_type, "보고서"),
        report_type=report_type,
        row=row,
    )

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _union_columns(rows: Iterable[Mapping[str, Any]]) -> list[str]:
    columns: list[str] = []
    seen: set[str] = set()

    for row in rows:
        for key in row.keys():
            if key not in seen:
                seen.add(key)
                columns.append(key)

    return columns


def _add_table_sheet(
    wb: Workbook,
    sheet_name: str,
    rows: list[Mapping[str, Any]],
) -> None:
    ws = wb.create_sheet(sheet_name)
    ws.freeze_panes = "A2"

    if not rows:
        ws["A1"] = "조건에 해당하는 데이터가 없습니다."
        return

    columns = _union_columns(rows)
    ws.append([REPORT_FIELD_LABELS.get(column, column) for column in columns])
    _style_header_row(ws, 1)

    for row in rows:
        ws.append([_display_value(row.get(column, "")) for column in columns])

    for row_cells in ws.iter_rows():
        for cell in row_cells:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    ws.auto_filter.ref = ws.dimensions
    _autosize_columns(ws)


def _add_link_status_sheet(
    wb: Workbook,
    linked_rows: list[Mapping[str, Any]],
) -> None:
    ws = wb.create_sheet("3종 연계 현황")
    ws.freeze_panes = "A2"

    headers = [
        "문서번호",
        "연도",
        "중심 격자 ID",
        "시도",
        "시군구",
        "발생 예측",
        "현장 예찰",
        "방제",
        "전체 연결 상태",
    ]
    ws.append(headers)
    _style_header_row(ws, 1)

    for item in linked_rows:
        ws.append([
            item.get("document_no", ""),
            item.get("year", ""),
            item.get("center_grid_id", ""),
            item.get("sido_name", ""),
            item.get("sigungu_name", ""),
            "연결 완료" if item.get("prediction_exists") else "미연결",
            "연결 완료" if item.get("field_survey_exists") else "미연결",
            "연결 완료" if item.get("control_exists") else "미연결",
            "3종 연결 완료" if item.get("fully_linked") else "일부 미연결",
        ])

    for row_cells in ws.iter_rows(min_row=2):
        for cell in row_cells:
            cell.alignment = Alignment(vertical="center")

    ws.auto_filter.ref = ws.dimensions
    _autosize_columns(ws)


def build_linked_reports_workbook(
    prediction_rows: list[Mapping[str, Any]],
    field_survey_rows: list[Mapping[str, Any]],
    control_rows: list[Mapping[str, Any]],
    linked_rows: list[Mapping[str, Any]],
) -> bytes:
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    _add_table_sheet(wb, "발생 예측", prediction_rows)
    _add_table_sheet(wb, "현장 예찰", field_survey_rows)
    _add_table_sheet(wb, "방제", control_rows)
    _add_link_status_sheet(wb, linked_rows)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
